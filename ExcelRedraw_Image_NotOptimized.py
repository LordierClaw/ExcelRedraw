from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

class CellData:
    def __init__(self, x, y, color):
        self.x = x
        self.y = y
        self.color = color

def imageProcess(imgInput, cellSize):
    img = imgInput.copy()
    imgWidth, imgHeight = img.size
    x = 0
    y = 0
    isNewColumn = False
    for m in range(0, imgWidth, cellSize):
        for n in range(0, imgHeight, cellSize):
            imgBox = img.crop((m, n, m+cellSize, n+cellSize))
            cellColor = getDominantColor(imgBox)
            if isNewColumn == True:
                x -= imgHeight/cellSize
                isNewColumn = False
            listOfCells.append(CellData(x, y, cellColor))
            x +=1
        y +=1
        isNewColumn = True

def getDominantColor(imgInput):
    img = imgInput.copy()
    img.convert("RGB")
    img.resize((1,1))
    rgb = img.getpixel((0,0))
    dominantColor = '{:02x}{:02x}{:02x}'.format(rgb[0], rgb[1], rgb[2]) #convert rgb to hex
    return dominantColor

# Prepare everything
wbook = Workbook()
fileName = "example.jpg"
cPixel = 5 #size of a cell

# Image process, and get the cell datas
listOfCells = []
openFile = Image.open(fileName)
imageProcess(openFile, 5)
wsheet = wbook.create_sheet()
for imgCell in listOfCells:
    ex_x = imgCell.x + 1
    ex_y = imgCell.y + 1
    if ex_x == 1:
        cd = wsheet.column_dimensions[get_column_letter(ex_y)]
        cd.width = 2.8
    exCell = wsheet.cell(row=(ex_x), column=(ex_y))
    exCell.fill = PatternFill(patternType='solid', fgColor=str(imgCell.color))
listOfCells.clear()
wbook.save("Final.xlsx")
print("Done.")  