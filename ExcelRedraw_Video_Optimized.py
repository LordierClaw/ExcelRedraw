from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.cell import WriteOnlyCell

class CellData:
    def __init__(self, x, y, color):
        self.x = x
        self.y = y
        self.color = color

def imageProcess(imgInput, cellSize):
    img = imgInput.copy()
    imgWidth, imgHeight = img.size
    x = 0
    y = 0
    isNewRow = False
    for iH in range(0, imgHeight, cellSize):
        for iW in range(0, imgWidth, cellSize):
            imgBox = img.crop((iW, iH, iW + cellSize, iH + cellSize))
            cellColor = get_Dominant_Color(imgBox)
            if isNewRow == True:
                y -= imgWidth/cellSize
                isNewRow = False
            listOfCells.append(CellData(x, y, cellColor))
            y +=1
        x +=1
        isNewRow = True

def get_Dominant_Color(imgInput):
    img = imgInput.copy()
    img.convert("RGB")
    img.resize((1,1))
    rgb = img.getpixel((0,0))
    dominantColor = '{:02x}{:02x}{:02x}'.format(rgb[0], rgb[1], rgb[2]) #convert rgb to hex
    return dominantColor

def get_number_of_rows_columns(firstFrame, cellSize):
    openFile = Image.open(firstFrame)
    imgWidth, imgHeight = openFile.size
    num_rows = int(imgWidth/cellSize)
    num_column = int(imgHeight/cellSize)
    return num_rows, num_column

#Prepare everything
wbookName = "Final.xlsx"
wbook = Workbook(write_only=True)
cPixel = 5 #size of a cell
num_frames = 2 #number of frames

# Image process, and get the cell datas
num_rows, num_columns = get_number_of_rows_columns("frame0000.jpg", cPixel)
listOfCells = []
for k in range(0, num_frames): # k is for the number of sheets
    fileName = "frame" + f"{k:04}" + ".jpg"
    openFile = Image.open(fileName)
    imageProcess(openFile, cPixel)
    wsheet = wbook.create_sheet(str(k))
    i = 0 # i is for the number of cells in listOfCells, which only increase in index
    for each_row in range(1, num_rows + 1):
        ex_row = []
        for each_column in range(1, num_columns + 1):
            ex_Cell = WriteOnlyCell(wsheet)
            ex_Cell.fill = PatternFill(patternType='solid', fgColor=str(listOfCells[i].color))
            ex_row.append(ex_Cell)
            i +=1
        wsheet.append(ex_row)
    print("Frame " + str(k+1) + "/" + str(num_frames) +" completed")
    listOfCells.clear()
print("Saving the final workbook file...")
wbook.save(wbookName)
print("Done.")  